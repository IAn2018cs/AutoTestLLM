# coding=utf-8
import re
from typing import Optional

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill, Alignment, Font

from utils.tools import create_path, root_relative_path


def truncate_string(s: str, max_length: int):
    # 如果字符串长度小于或等于31，则返回全部字符串
    if len(s) <= max_length:
        return s
    # 否则，截取前31个字符
    else:
        return s[:max_length]


def remove_parentheses_content_fixed(text):
    # 更新正则表达式模式以确保处理中文字符
    pattern = r'\（.*?\）'
    # 替换括号及其内容为空字符串
    cleaned_text = re.sub(pattern, '', text).strip()
    if "：" in cleaned_text:
        return cleaned_text.split("：")[0].strip()
    if ":" in cleaned_text:
        return cleaned_text.split(":")[0].strip()
    return cleaned_text


def get_worksheet_name(name: str) -> str:
    cleaned_text = remove_parentheses_content_fixed(name)

    # 替换所有Excel不支持的字符
    invalid_chars = "[]:*?/\\"
    for char in invalid_chars:
        cleaned_text = cleaned_text.replace(char, "_")
    return truncate_string(cleaned_text, 30)


def clean_illegal_chars(data_frame):
    """ 使用 openpyxl 的正则表达式清理 DataFrame 中的非法字符。"""
    for col in data_frame.columns:
        if data_frame[col].dtype == object:  # 只处理字符串列
            data_frame[col] = data_frame[col].apply(
                lambda x: ILLEGAL_CHARACTERS_RE.sub('', x) if isinstance(x, str) else x
            )
    return data_frame


def create_work(name, data_frame, ex_writer):
    df = clean_illegal_chars(pd.DataFrame(data_frame))
    sheet_name = get_worksheet_name(name)
    df.to_excel(ex_writer, sheet_name=sheet_name, index=False)

    # 由于openpyxl的Workbook是在保存后才创建的，需要手动获取并修改Workbook对象
    workbook = ex_writer.book
    # 获取工作表
    worksheet = workbook[sheet_name]
    # 设置第一列的宽度为 ，其他列宽度为
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 8
    for col in worksheet.iter_cols(min_col=3, max_col=worksheet.max_column):
        worksheet.column_dimensions[col[0].column_letter].width = 30
    # 固定第一行
    worksheet.freeze_panes = 'A2'
    # 设置所有单元格的自动换行
    wrap_text_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = wrap_text_alignment
    # 设置第一行特定的格式
    for cell in worksheet[1]:
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = wrap_text_alignment
    worksheet.row_dimensions[1].height = 30
    # 设置其他行的行高自适应（简单估算）
    approx_char_per_line = 24  # 每行大约字符数
    min_height = 30  # 最小行高
    line_height = 15
    for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        max_char_in_row = max((len(str(cell.value)) for cell in row), default=0)
        lines_needed = max(1, int(max_char_in_row / approx_char_per_line))  # 确保至少为1行
        worksheet.row_dimensions[idx].height = max(min_height, line_height * lines_needed)


def create_worksheet(file_name: str, map_data: dict) -> Optional[str]:
    temp_path = root_relative_path('temp_files')
    create_path(temp_path)
    path = f'{temp_path}/{file_name}.xlsx'
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for key, value in map_data.items():
            if len(value) != 0:
                create_work(key, value, writer)
    return path


def create_style_excel(file_name: str, sheet_name: str, data_frame: list[dict]) -> str:
    temp_path = root_relative_path('temp_files')
    create_path(temp_path)
    path = f'{temp_path}/{file_name}.xlsx'
    with pd.ExcelWriter(path, engine='openpyxl') as ex_writer:
        create_work(sheet_name, data_frame, ex_writer)
    return path
